# -*- coding:utf-8 -*-
import os

from flask import abort, current_app, send_from_directory
from flask import request
from werkzeug.datastructures import MultiDict

from api.lib.common_setting.employee import EmployeeCRUD, EmployeeAddForm, EmployeeUpdateByUidForm
from api.lib.common_setting.resp_format import ErrFormat
from api.resource import APIView

prefix = '/employee'


class EmployeeView(APIView):
    url_prefix = (f'{prefix}',)

    def get(self):
        department_id = int(request.args.get('department_id', 0))
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 10))
        search = request.args.get('search', '')
        order = request.args.get('order', '')
        block_status = int(request.args.get('block_status', -1))

        employee_list = EmployeeCRUD.get_employee_list_by(
            department_id, block_status, search, order, page, page_size)

        return self.jsonify(employee_list)

    def post(self):
        form = EmployeeAddForm(MultiDict(request.json))
        if not form.validate():
            abort(400, ','.join(['{}: {}'.format(filed, ','.join(msg))
                                 for filed, msg in form.errors.items()]))

        data = EmployeeCRUD.add(**form.data)
        return self.jsonify(data.to_dict())


class EmployeeFilterView(APIView):
    url_prefix = (f'{prefix}/filter',)

    def post(self):
        params = request.json
        department_id = int(params.get('department_id', 0))
        page = int(params.get('page', 1))
        page_size = int(params.get('page_size', 10))
        search = params.get('search', '')
        order = params.get('order', '')
        block_status = int(params.get('block_status', -1))
        conditions = list(params.get("conditions", []))
        employee_list = EmployeeCRUD.get_employee_list_by_body(department_id, block_status, search, order, conditions,
                                                               page, page_size)

        return self.jsonify(employee_list)


class EmployeeViewWithId(APIView):
    url_prefix = (f'{prefix}/<int:_id>',)

    def get(self, _id):
        data = EmployeeCRUD.get_employee_by_id(_id)
        return self.jsonify(data.to_dict())

    def put(self, _id):
        params = request.json
        direct_supervisor_id = params.get('direct_supervisor_id', None)
        if direct_supervisor_id and int(_id) == int(direct_supervisor_id):
            abort(400, ErrFormat.direct_supervisor_is_not_self)

        data = EmployeeCRUD.update(_id, **params)
        return self.jsonify(data.to_dict())


class EmployeeCountView(APIView):
    url_prefix = (f'{prefix}/count',)

    def get(self):
        block_status = int(request.args.get('block_status', -1))
        employee_count = EmployeeCRUD.get_employee_count(block_status)
        return self.jsonify(employee_count=employee_count)


class EmployeeImportView(APIView):
    url_prefix = (f'{prefix}/import',)

    def post(self):
        employee_list = request.json.get('employee_list', [])
        if not employee_list:
            abort(400, ErrFormat.employee_list_is_empty)
        result = EmployeeCRUD.import_employee(employee_list)
        return self.jsonify(result)


class EmployeeBatchView(APIView):
    url_prefix = (f'{prefix}/batch',)

    def post(self):
        params = request.json
        column_name = params.get('column_name', None)
        employee_id_list = params.get('employee_id_list', None)
        column_value = params.get('column_value', None)
        if column_name not in ['department_id', 'direct_supervisor_id', 'position_name', 'password', 'block']:
            abort(400, ErrFormat.column_name_not_support)
        result = EmployeeCRUD.batch_employee(
            column_name, column_value, employee_id_list)
        return self.jsonify(result)


class EmployeeViewWithACLID(APIView):
    url_prefix = (f'{prefix}/by_uid/<int:_uid>',)

    def get(self, _uid):
        result = EmployeeCRUD.get_employee_by_uid_with_create(_uid)
        return self.jsonify(result)

    def put(self, _uid):
        form = EmployeeUpdateByUidForm(MultiDict(request.json))
        if not form.validate():
            abort(400, ','.join(['{}: {}'.format(filed, ','.join(msg))
                                 for filed, msg in form.errors.items()]))

        data = EmployeeCRUD.edit_employee_by_uid(_uid, **form.data)
        return self.jsonify(data.to_dict())


class EmployeeChangePasswordWithACLID(APIView):
    url_prefix = (f'{prefix}/by_uid/change_password/<int:_uid>',)

    def put(self, _uid):
        password = request.json.get('password', None)
        if not password:
            abort(400, ErrFormat.password_is_required)

        data = EmployeeCRUD.change_password_by_uid(_uid, password)
        return self.jsonify(200)


class EmployeePositionView(APIView):
    url_prefix = (f'{prefix}/position',)

    def get(self):
        """"""
        result = EmployeeCRUD.get_all_position()
        return self.jsonify(result)


class EmployeeViewExportExcel(APIView):
    url_prefix = (f'{prefix}/export_all',)

    def get(self):
        excel_filename = 'all_employee_info.xlsx'
        excel_path = current_app.config['UPLOAD_DIRECTORY_FULL']
        excel_path_with_filename = os.path.join(excel_path, excel_filename)

        block_status = int(request.args.get('block_status', -1))
        data_list = EmployeeCRUD.get_export_employee_list(block_status)

        headers = data_list[0].keys()
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        # insert header
        for col_num, col_data in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=col_data)

        for row_num, row_data in enumerate(data_list, start=2):
            for col_num, col_data in enumerate(row_data.values(), start=1):
                ws.cell(row=row_num, column=col_num, value=col_data)
        wb.save(excel_path_with_filename)
        return send_from_directory(excel_path, excel_filename, as_attachment=True)
